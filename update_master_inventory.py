"""
update_master_inventory.py
=============================
Último paso del pipeline ETL: toma el inventario fusionado
(merged_inventory.csv, salida de merge_inventories.py) y lo
deposita sobre una COPIA del spreadsheet maestro (.ods/.xlsx),
sin modificar jamás el archivo maestro original.

El script:
1. Crea una copia física del spreadsheet maestro.
2. Localiza, dentro del maestro, la fila que corresponde al header
   real, buscando la primera fila cuyo contenido (en el mismo
   orden) coincida EXACTAMENTE con el header de
   merged_inventory.csv. Esto valida simultáneamente que los
   nombres de columna y sus posiciones sean idénticos entre ambos
   archivos, sin necesitar rundeck_header_list.txt.
3. Valida que la cantidad de filas de datos del maestro (a partir
   del header detectado) coincida exactamente con la cantidad de
   filas de merged_inventory.csv. Si no coincide, aborta: no es
   seguro asumir correspondencia posicional entre ambos archivos.
4. Compara celda a celda (con normalización numérica, para no
   reportar diferencias de formato como "16" vs "16.0" como si
   fueran cambios reales) y sobrescribe únicamente las celdas cuyo
   valor cambió.
5. Informa qué columnas cambiaron y, para cada una, qué celdas
   específicas fueron modificadas.

Mecanismo de edición:
- .xlsx: se copia el archivo y se edita directamente con openpyxl,
  preservando el formato original.
- .ods: dado que editar celdas individuales de un .ods de forma
  robusta requiere manejar manualmente la compresión de celdas/filas
  repetidas del formato ODF (atributos number-columns-repeated /
  number-rows-repeated), y el pipeline ya depende de LibreOffice
  headless, se usa un round-trip: .ods -> .xlsx (LibreOffice) ->
  edición con openpyxl -> .xlsx -> .ods (LibreOffice).

El proceso completo es atómico: se trabaja sobre un archivo
temporal en el mismo directorio de salida, y solo se reemplaza el
archivo final mediante os.replace() si todo el proceso terminó sin
errores.

Uso:
    python3 update_master_inventory.py \
        merged_inventory.csv \
        master_inventory.(ods|xlsx) \
        [master_inventory_updated.(ods|xlsx)]
"""

import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from base_inventory import error, strip_quotes, usage
from merge_inventories import load_csv_data, validate_rows
from prepare_master_inventory import find_libreoffice

SUPPORTED_EXTENSIONS = {".ods", ".xlsx"}


# ============================================================
# NORMALIZACIÓN Y CONVERSIÓN DE VALORES
# ============================================================

def normalize_value(value) -> str:
    """
    Normaliza un valor (proveniente de una celda del spreadsheet o
    de un campo del CSV) a una representación canónica en string,
    para poder comparar equivalencia sin falsos positivos por
    diferencias de formato numérico (ej. 16 vs 16.0).
    """
    if value is None:
        return ""
    text = str(value).strip()
    try:
        as_float = float(text)
    except ValueError:
        return text
    if as_float == int(as_float):
        return str(int(as_float))
    return str(as_float)


def coerce_for_cell(value_str: str) -> int | float | str | None:
    """
    Convierte un valor string proveniente del CSV al tipo Python
    más apropiado para escribirlo en una celda, preservando el
    tipo numérico cuando corresponde (en vez de forzar todo a
    texto, lo que degradaría el formato de columnas numéricas).
    """
    if value_str == "":
        return None
    try:
        return int(value_str)
    except ValueError:
        pass
    try:
        return float(value_str)
    except ValueError:
        pass
    return value_str


def column_letter(index0: int) -> str:
    """
    Convierte un índice de columna 0-based a notación de letras
    estilo spreadsheet (0 -> A, 25 -> Z, 26 -> AA, ...).
    """
    index = index0 + 1
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


# ============================================================
# LIBREOFFICE: CONVERSIÓN GENÉRICA ENTRE FORMATOS
# ============================================================

def convert_with_libreoffice(
    libreoffice: str,
    input_path: Path,
    target_format: str,
    out_dir: Path,
) -> Path:
    command = [
        libreoffice,
        "--headless",
        "--convert-to", target_format,
        "--outdir", str(out_dir),
        str(input_path),
    ]
    result = subprocess.run(
        command,
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        error(
            f"LibreOffice no pudo convertir el archivo a "
            f"{target_format}.\n\n"
            f"stdout:\n{result.stdout}\n\n"
            f"stderr:\n{result.stderr}"
        )
    converted = out_dir / f"{input_path.stem}.{target_format}"
    if not converted.is_file():
        error(
            "LibreOffice terminó sin error, pero no se encontró "
            f"el archivo convertido:\n  {converted}"
        )
    return converted


# ============================================================
# LOCALIZAR HEADER Y CONTAR FILAS DE DATOS EN EL MAESTRO
# ============================================================

def find_header_row_number(ws: Worksheet, merged_header: list[str]) -> int | None:
    """
    Busca la primera fila cuyo contenido, en el mismo orden,
    coincida EXACTAMENTE con merged_header. Esto valida a la vez
    nombres y posiciones de columna entre el maestro y el
    inventario fusionado.
    """
    n_cols = len(merged_header)
    for row_number, row in enumerate(
        ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            max_col=n_cols,
            values_only=True,
        ),
        start=1,
    ):
        candidate = ["" if v is None else str(v).strip() for v in row]
        if candidate == merged_header:
            return row_number
    return None


def count_master_data_rows(ws: Worksheet, header_row_number: int, n_cols: int) -> int:
    """
    Cuenta filas de datos a partir de header_row_number + 1,
    considerando como límite la última fila con al menos un valor
    no vacío dentro de las n_cols columnas relevantes.
    """
    last_row_with_data = header_row_number
    for row_number, row in enumerate(
        ws.iter_rows(
            min_row=header_row_number + 1,
            max_row=ws.max_row,
            max_col=n_cols,
            values_only=True,
        ),
        start=header_row_number + 1,
    ):
        if any(v is not None and str(v).strip() != "" for v in row):
            last_row_with_data = row_number
    return last_row_with_data - header_row_number


# ============================================================
# ACTUALIZAR UN ARCHIVO XLSX (EDICIÓN DIRECTA)
# ============================================================

def get_active_worksheet(wb: Workbook) -> Worksheet:
    """Obtiene la hoja de cálculo activa de un libro de trabajo."""
    ws = wb.active

    if not isinstance(ws, Worksheet):
        error(
            "El spreadsheet no contiene una hoja de cálculo activa "
            "válida para actualizar."
        )

    return ws


def get_cell(ws: Worksheet, row: int, column: int) -> Cell:
    """Obtiene una celda normal y rechaza celdas combinadas."""
    cell = ws.cell(row=row, column=column)

    if isinstance(cell, MergedCell):
        error(
            "No se puede actualizar una celda combinada "
            f"en {column_letter(column - 1)}{row}."
        )

    return cell


def update_xlsx(
    path: Path,
    merged_header: list[str],
    merged_rows: list[list[str]],
) -> dict[str, list[str]]:
    """
    Abre el .xlsx en `path`, localiza el header, valida
    correspondencia de filas, sobrescribe únicamente las celdas
    cuyo valor cambió, guarda el archivo y retorna un mapa
    {nombre_columna: [direcciones_de_celda_modificadas]}.
    """
    wb = openpyxl.load_workbook(path)
    ws = get_active_worksheet(wb)

    n_cols = len(merged_header)
    header_row_number = find_header_row_number(ws, merged_header)
    if header_row_number is None:
        error(
            "No se encontró ninguna fila en el spreadsheet maestro "
            "que coincida exactamente (mismos nombres y mismas "
            "posiciones) con el header del inventario fusionado."
        )

    print(f"Fila de header detectada en el maestro : {header_row_number}")

    master_data_rows = count_master_data_rows(
        ws, header_row_number, n_cols
    )

    if master_data_rows != len(merged_rows):
        error(
            "La cantidad de filas de datos del spreadsheet maestro "
            f"({master_data_rows}) no coincide con la cantidad de "
            f"filas del inventario fusionado ({len(merged_rows)}).\n"
            "No es seguro asumir correspondencia posicional entre "
            "ambos archivos."
        )

    changes: dict[str, list[str]] = {}
    first_data_row = header_row_number + 1

    for row_offset, merged_row in enumerate(merged_rows):
        sheet_row_number = first_data_row + row_offset
        for col_idx in range(n_cols):
            cell = get_cell(ws, sheet_row_number, col_idx + 1)
            current_norm = normalize_value(cell.value)
            new_value_str = merged_row[col_idx]
            new_norm = normalize_value(new_value_str)

            if current_norm == new_norm:
                continue    # No hay cambio, no sobrescribir.

            cell.value = coerce_for_cell(new_value_str)

            column_name = merged_header[col_idx]
            address = f"{column_letter(col_idx)}{sheet_row_number}"
            changes.setdefault(column_name, []).append(address)

    wb.save(path)
    return changes


# ============================================================
# REPORTE DE CAMBIOS
# ============================================================

def print_change_report(
    changes: dict[str, list[str]],
    merged_header: list[str],
) -> None:
    print()
    if not changes:
        print(
            "No se detectaron cambios respecto al spreadsheet "
            "maestro previo."
        )
        return

    print("Columnas modificadas:")
    print()
    for col_idx, column_name in enumerate(merged_header):
        if column_name not in changes:
            continue
        letter = column_letter(col_idx)
        addresses = changes[column_name]
        print(f"Columna {letter} ({column_name})")
        print(", ".join(addresses))
        print()


# ============================================================
# MAIN
# ============================================================

def main() -> None:
    if len(sys.argv) < 3:
        usage(
            f"Uso: {sys.argv[0]} "
            "merged_inventory.csv "
            "master_inventory.(ods|xlsx) "
            "[master_inventory_updated.(ods|xlsx)]"
        )

    if len(sys.argv) > 4:
        usage(
            f"Cantidad de argumentos inválida.\n"
            f"Uso: {sys.argv[0]} "
            "merged_inventory.csv "
            "master_inventory.(ods|xlsx) "
            "[master_inventory_updated.(ods|xlsx)]"
        )

    merged_csv_path = Path(sys.argv[1])
    master_path = Path(sys.argv[2])

    for path, label in (
        (merged_csv_path, "inventario fusionado"),
        (master_path, "spreadsheet maestro"),
    ):
        if not path.is_file():
            error(f"No se encontró el {label}:\n  {path}")

    ext = master_path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        error(
            "El spreadsheet maestro debe ser un archivo ODS "
            "(.ods) o Excel (.xlsx):\n"
            f"  {master_path}"
        )

    output_path = (
        Path(sys.argv[3])
        if len(sys.argv) >= 4
        else master_path.parent
        / f"{master_path.stem}_updated{master_path.suffix}"
    )

    if output_path.suffix.lower() != ext:
        error(
            "La extensión del archivo de salida debe coincidir "
            "con la del spreadsheet maestro "
            f"({ext}):\n  {output_path}"
        )

    # --------------------------------------------------------
    # Validar colisiones entre archivos de entrada y salida.
    # --------------------------------------------------------
    resolved_output = output_path.resolve()
    resolved_merged = merged_csv_path.resolve()
    resolved_master = master_path.resolve()

    if resolved_output == resolved_merged:
        error(
            "El archivo de salida no puede ser el mismo archivo "
            "que el inventario fusionado:\n"
            f"  {output_path}"
        )
    if resolved_output == resolved_master:
        error(
            "El archivo de salida no puede ser el mismo archivo "
            "que el spreadsheet maestro (se debe generar una "
            "copia):\n"
            f"  {output_path}"
        )

    # --------------------------------------------------------
    # Cargar inventario fusionado.
    # --------------------------------------------------------
    _header_line, header_fields_raw, rows_raw = load_csv_data(
        merged_csv_path
    )
    merged_header = [strip_quotes(f) for f in header_fields_raw]
    n_cols = len(merged_header)

    rows_raw = validate_rows(rows_raw, n_cols, merged_csv_path)
    merged_rows = [[strip_quotes(v) for v in row] for row in rows_raw]

    print()
    print(f"Inventario fusionado             : {merged_csv_path}")
    print(f"Spreadsheet maestro               : {master_path}")
    print(f"Spreadsheet de salida             : {output_path}")
    print(f"Columnas en inventario fusionado : {n_cols}")
    print(f"Filas de datos en inventario fusionado : {len(merged_rows)}")

    # --------------------------------------------------------
    # Procesar: crear copia + editar + reemplazo atómico.
    # --------------------------------------------------------
    with tempfile.TemporaryDirectory(prefix="update_master_") as tmp_dir_str:
        tmp_dir = Path(tmp_dir_str)

        with tempfile.NamedTemporaryFile(
            dir=output_path.parent,
            prefix=f".{output_path.name}.",
            suffix=".tmp",
            delete=False,
        ) as tmp_out_f:
            temporary_output_path = Path(tmp_out_f.name)

        try:
            if ext == ".xlsx":
                shutil.copy2(master_path, temporary_output_path)
                changes = update_xlsx(
                    temporary_output_path,
                    merged_header,
                    merged_rows,
                )
            else:
                libreoffice = find_libreoffice()

                print()
                print(
                    "Convirtiendo copia del maestro a XLSX "
                    "(edición intermedia)..."
                )
                xlsx_tmp = convert_with_libreoffice(
                    libreoffice, master_path, "xlsx", tmp_dir
                )

                changes = update_xlsx(
                    xlsx_tmp,
                    merged_header,
                    merged_rows,
                )

                print("Reconvirtiendo copia editada a ODS...")
                ods_tmp = convert_with_libreoffice(
                    libreoffice, xlsx_tmp, "ods", tmp_dir
                )

                shutil.copy2(ods_tmp, temporary_output_path)

            os.replace(temporary_output_path, output_path)
            temporary_output_path = None
        finally:
            if temporary_output_path is not None:
                try:
                    temporary_output_path.unlink()
                except FileNotFoundError:
                    pass

    # --------------------------------------------------------
    # Reporte de cambios.
    # --------------------------------------------------------
    print_change_report(changes, merged_header)

    # --------------------------------------------------------
    # Resultado.
    # --------------------------------------------------------
    print()
    print("[OK] Spreadsheet maestro actualizado correctamente.")
    print(f"Inventario fusionado : {merged_csv_path}")
    print(f"Maestro original     : {master_path}")
    print(f"Maestro actualizado  : {output_path}")


if __name__ == "__main__":
    main()